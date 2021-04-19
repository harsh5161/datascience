import pandas as pd
import numpy as np
import xlrd
import csv
import collections 
<<<<<<< HEAD:segmentation/seg_userInputs.py

=======
from openpyxl import load_workbook
>>>>>>> master:predictive/userInputs.py
def importFile(path,nrows=None):

    print('#### RUNNING WAIT ####')

    # IF THE EXTENSION IS CSV
    def importCsv(path):

        print('We have a csv file')
        try:
            df = pd.read_csv(path,low_memory=False,nrows=nrows,error_bad_lines=False)
            if df.shape[1] == 1:
                df = pd.read_csv(path,low_memory=False,sep=';',nrows=nrows)
            print('This file has {} columns and {} rows'.format(df.shape[1],df.shape[0]))
            return df

        except FileNotFoundError:
            print('File not found, Check the name, path, spelling mistakes')
            error = True
            return None

        except UnicodeDecodeError:
            try:
                enc = 'unicode_escape'
                df = pd.read_csv(path,encoding=enc,low_memory=False,nrows=nrows,error_bad_lines=False)
                print('This file has {} columns and {} rows'.format(df.shape[1],df.shape[0]))
                return df

            except UnicodeDecodeError:
                try:
                    enc = 'ISO-8859-1'
                    df = pd.read_csv(path,encoding=enc,low_memory=False,nrows=nrows,error_bad_lines=False)
                    print('This file has {} columns and {} rows'.format(df.shape[1],df.shape[0]))
                    return df
                except:
                    pass

        except:
            try:
                df= pd.read_csv(path,nrows=nrows,error_bad_lines=False)
                separators= ["~","!","@","#","$","%","^","&","*",":","|","/",";"]     # all possible separators
                if len(df.columns)<=3 :                                               # if separator was "," we would have more than 1 columns
                    cols = df.columns[0]
                    possibleSep = []
                    for i in separators:                                    # checking all the separators present in column names
                        if i in cols:
                            possibleSep.append(i)

                    for j in possibleSep:                                   # iterate through possible seprators till we get the correct one
                        df_sep = pd.read_csv(path,sep=j,nrows=nrows,error_bad_lines=False)
                        if len(df_sep.columns)>3:
                            print('This file has {} columns and {} rows'.format(df_sep.shape[1],df_sep.shape[0]))
                            return df_sep
            except:
                try:
                    if len(pd.read_csv(path,sep=None).columns,nrows=nrows,error_bad_lines=False)>3  :                   # for tab ie "\" tsv files
                        df = pd.read_csv(path,sep=None,nrows=nrows,error_bad_lines=False)
                        print('This file has {} columns and {} rows'.format(df.shape[1],df.shape[0]))
                        return df
                except:
                    pass

    # IF THE EXTENSION IS JSON
    def importJSON(path):
        try:
            print('We have a JSON file')
            df = pd.read_json(path)
            print('This file has {} columns and {} rows'.format(df.shape[1],df.shape[0]))
            return df
        except Exception:
            try:
                df = pd.read_json(path,lines=True)
                print('This file has {} columns and {} rows'.format(df.shape[1],df.shape[0]))
                return df

            except ValueError:
                print('File not found, Check the name, path, spelling mistakes')
                error = True
                return None

    # IF THE EXTENSION IS XL
    def importExcel(path):
        try:
            print('We have an Excel file')
            #######
            # opening workbook
            wb = load_workbook(path)

            if len(wb.sheetnames) == 1 :
                data = wb[wb.sheetnames[0]].values
                cols = next(data)[0:]
            else:
                val = input('Input the name of the sheet')
                data = wb[val].values
                cols =  next(data)[0:]
                
            df = pd.DataFrame(data,columns=cols)

            return df

            #######
        except FileNotFoundError:
            print('File not found, Check the name, path, spelling mistakes')
            error = True
            return None

    def importTable(path):
        try:
            print('We have General Table File')
            df = pd.read_table(path,nrows=nrows)
            if df.shape[1] == 1:
                df = pd.read_table(path,sep=',',nrows=nrows)
            print('This file has {} columns and {} rows'.format(df.shape[1],df.shape[0]))
            return df
        except FileNotFoundError:
            print('File not found, Check the name, path, spelling mistakes')
            error = True
            return None

    try:
        ext = path.split('.')[1].lower().strip()
        print('extension is {}'.format(ext))
        if ext == 'csv' or ext == 'tsv':
            df = importCsv(path)
            df = duplicateHandler(df)
            return df,None
        elif ext == 'json':
            df = importJSON(path)
            df = duplicateHandler(df)
            return df,None
        elif 'xl' in ext:
            df = importExcel(path)
            return df,None
        elif ext == 'data':
            df = importTable(path)
            df = duplicateHandler(df)
            return df,None
        else:
            print('File format not supported\n')
    except Exception as e:
        print('We ran into some Error!')
        print('The error message is {}'.format(e))
        return None,None

######################################
######## getUserInput Function #######
######################################

def getUserInput(df):
    if isinstance(df,pd.DataFrame):
        print('\nDataFrame Succesfully imported\n')
        print(df.columns)
<<<<<<< HEAD:segmentation/seg_userInputs.py
        # Get Key Column
        key = getKey(df.columns)
        if not key:
            key = findKey(df.columns[0])
        if key:
            print("Dropping Key from the axis")
            df.drop(key,axis=1,inplace=True)

        # Remove User Specified ID Columns
        cluster_list,profile_list = removeUserSpecifiedIDs(df)
        print('Cluster List',cluster_list)
        print('Profile List',profile_list)
        info = {'key':key,'cols':cluster_list,'profile_cols':profile_list}
=======

        # Get Target from user
        target = getTarget(df.columns)

        if not target:
            # Quit the whole process
            print('\nQuitting Process\n')
            return None
        else:
            # Get Key Column
            key = getKey(df.columns)
            if not key:
                key = findKey(df.columns[0])
            if key:
                df.drop(key,axis=1,inplace=True)

            # Remove User Specified ID Columns
            df = removeUserSpecifiedIDs(df,True)

            # Remove Successive Targets
            df = removeUserSpecifiedIDs(df)

            # Quick/Slow results for max evals
            quick = quick_slow()
            if quick:print('QUICK MODELLING WITH DEFAULT PARAMETERS')
            else:print('HyperOP with MAX EVALS = 15')

            graph = disable_graphs()
            if graph:print("Graphs are now turned off for this output session")
            else:print("Graphs are now turned on for this output session")

        info = {'target':target,'key':key,'cols':df.drop([target],axis=1).columns.to_list(),'q_s':quick,'graph':graph}
>>>>>>> master:predictive/userInputs.py

        return info
    else:
        return None

######################################
######## getKey Function #######
######################################

def getKey(columns):
    print('\nEnter \'quit\' to quit')
    key = input('Enter the Key/Identification Column : ')
    if key == 'quit':
        return None,False
    elif key in columns.values:
        print('Key Spotted!')
        return key
    else:
        print('Key {} Not found in the data'.format(key))
        print('Preview can\'t be shown!!')
        return None

######################################
######## findKey Function #######
######################################

def findKey(column):
    if 'id' in column.lower():
        dec = input("Is the column \'{}\' an identification column? If yes, enter y : ".format(column))
        if dec == 'y':
            print(f'Identification column {column} obtained')
            return column
        else:
            print('Identification column not obtained/found')
            return None

######################################
######## removeUserSpecifiedIDs Function #######
######################################
def Splitter(cols,full_list):
    full_list = full_list[:]
    removed_cols = []
    not_found_cols = []
    if not cols:
        print('No Columns removed')
        return full_list
    else:
        try:
            columns  = cols.split(',')
            for column in columns:
                if column in full_list:
                    full_list.remove(column)
                    removed_cols.append(column)
                else:
                    not_found_cols.append(column)
            if removed_cols:
                print('\n{} columns are removed as entered by the user'.format(len(removed_cols)))
            if not_found_cols:
                print('\n{}'.format(not_found_cols))
                print('These columns were not found, hence not removed')
            return full_list
        except:
            print('Invalid Entry of columns! No Columns removed')
            return full_list
def removeUserSpecifiedIDs(df):
    full_list = df.columns.to_list()
    cluster_list = []
    profile_list = []
    cols = input("Would you like to remove any columns that you dont want to be included in clustering, if so enter the columns in this format separated by commas: col1,col2,col3 ")
    cluster_list = Splitter(cols,full_list)
    cols = input("Would you like to remove any columns that you dont want to be included in the cluster profiling, if so enter the columns in this format separated by commas: col1,col2,col3 ")
    profile_list = Splitter(cols,full_list)
    return cluster_list,profile_list


######################################
######## dataHandler Function #######
######################################

######################################
######## disable_graphs Function #######
######################################

def disable_graphs():
    val = input("Do you want to disable graphs for this output session? If disable press y : ").lower()
    return True if val == 'y' else False

######################################
######## disable_graphs Function #######
######################################
def removeCommas(x):
    if ',' in str(x):
        try:
            return pd.to_numeric(str(x).replace(',',''))
        except:
            return x
    else:
        return x

def dataHandler(dx,target=None):
        update=False
        for col in dx.columns: 			#Counting the non-null values present in a column and removing them if necessary
            if 'Unnamed' in col:
                if dx[col].count()<0.5*dx.shape[0]:
                    dx.drop(col,axis=1,inplace=True)
                    update = True

        #checking if there are string specified null values in the target column and replacing it 
        if target is not None:
            if str(dx[target].dtype) == 'object':
                dx[target].replace({'NA':np.nan},inplace=True)
                if dx[target].nunique()>5:
                    dx[target] = dx[target].apply(lambda x: removeCommas(x))
                    try:
                        dx[target] = pd.to_numeric(dx[target])
                        if str(dx[target].dtype) != 'object':
                            dx[target] = dx[target].astype(float)
                    except Exception as e:
                        print(f'Exception has occurred : {e}')
                

        # to handel cases when some blank rows or other information above the data table gets assumed to be column name
        if (len([col for col in dx.columns if 'Unnamed' in col]) > 0.5*dx.shape[1]  ):#Checking for unnamed columns
            colNew = dx.loc[0].values.tolist()           # Getting the values in the first row of the dataframe into a list
            dx.columns = colNew                          #Making values stored in colNew as the new column names
            dx = dx.drop(labels=[0])                     #dropping the row whose values we made as the column names
            dx.reset_index(drop=True, inplace=True)      #resetting index to the normal pattern 0,1,2,3...
        else:
            return dx,update

        new_column_names=dx.columns.values.tolist() # Following three lines of code are for counting the number of null values in our new set of column names
        new_column_names=pd.DataFrame(new_column_names)
        null_value_sum=new_column_names.isnull().sum()[0]
        if null_value_sum<0.5*dx.shape[1]: # if count of null values are less than a certain ratio of total no of columns
            return dx,update
        while(null_value_sum>=0.5*dx.shape[1]):
            colNew = dx.loc[0].values.tolist()
            dx.columns = colNew
            dx = dx.drop(labels=[0])
            dx.reset_index(drop=True, inplace=True)
            new_column_names=dx.columns.values.tolist()
            new_column_names=pd.DataFrame(new_column_names)
            null_value_sum=new_column_names.isnull().sum()[0]
<<<<<<< HEAD:segmentation/seg_userInputs.py
        return dx


=======


        return dx,update

>>>>>>> master:predictive/userInputs.py
######################################
######## duplicateHandler Function #######
######################################

def duplicateHandler(df):
<<<<<<< HEAD:segmentation/seg_userInputs.py
=======

    df.rename(columns = {"" : 'Unnamed'}, inplace = True) #dealing with column names that are empty strings
>>>>>>> master:predictive/userInputs.py
    actual = df.columns.to_list()
    a = [x.strip().lower() for x in df.columns.to_list()]
    dups = [item for item, count in collections.Counter(a).items() if count > 1]

    for i in range(len(a)):
        if a[i] in dups:
            actual[i] = f'{actual[i].strip()}_{i}'
    
    df.columns = actual 
<<<<<<< HEAD:segmentation/seg_userInputs.py
    return df 
######################################
######## duplicateHandler Function #######
######################################
=======
    
    return df 
######################################
######## duplicateHandler Function #######
######################################
>>>>>>> master:predictive/userInputs.py
